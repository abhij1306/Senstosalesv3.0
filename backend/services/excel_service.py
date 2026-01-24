"""
Excel Generation Service
Uses XlsxWriter to generate formatted Excel reports with strict layout control.
"""

import io
import logging
import os
import sqlite3
from copy import copy

import xlsxwriter
from fastapi.responses import StreamingResponse

from backend.core.num_to_words import amount_to_words

logger = logging.getLogger(__name__)


class ExcelService:
    @staticmethod
    def generate_response(data: list[dict], report_type: str) -> StreamingResponse:
        """
        Convert list of dicts to Excel download response (Legacy fallback)
        """
        filename = f"{report_type}.xlsx"
        return ExcelService.generate_from_list(data, filename)

    @staticmethod
    def _write_standard_header(
        worksheet,
        workbook,
        columns: int,
        db: sqlite3.Connection,
        title: str = None,
        layout: str = "invoice",
        font_name: str = "Calibri",
    ):
        """
        Consistently writes the business header across all reports with layout options.
        layout: 'invoice' (Standard for Sales Invoice) or 'challan' (Standard for DC, Summary, GC)
        """
        # Fetch settings from DB
        try:
            rows = db.execute("SELECT key, value FROM settings").fetchall()
            settings = {row["key"]: row["value"] for row in rows}
        except Exception as e:
            logger.error(f"Failed to fetch business settings, using defaults: {e}")
            settings = {}

        # Default Fallbacks
        # Default Fallbacks - STRICTLY from Settings
        s_name = settings.get("supplier_name", "")
        s_desc = settings.get("supplier_description", "")
        s_addr = settings.get("supplier_address", "")
        s_gst = settings.get("supplier_gstin", "")
        s_phone = settings.get("supplier_contact", "")
        s_state = settings.get("supplier_state", "")
        s_state_code = settings.get("supplier_state_code", "")

        # Formats
        title_fmt = workbook.add_format({"bold": True, "font_size": 18, "align": "center", "font_name": font_name})
        subtitle_fmt = workbook.add_format({"bold": True, "font_size": 10, "align": "center", "font_name": font_name})
        tel_fmt = workbook.add_format({"font_size": 10, "align": "center", "font_name": font_name})
        name_fmt = workbook.add_format({"bold": True, "font_size": 14, "align": "left", "font_name": font_name})
        detail_fmt = workbook.add_format({"font_size": 11, "align": "left", "font_name": font_name})
        bold_detail = workbook.add_format({"bold": True, "font_size": 11, "align": "left", "font_name": font_name})

        def set_val(coord, val, fmt=detail_fmt, align=None):
            if val:
                worksheet.write(coord, val, fmt)

        row = 0

        if layout == "invoice":
            # Layout matching 'GST_INV_11.xls'
            if title:
                worksheet.merge_range(
                    row,
                    0,
                    row,
                    columns - 1,
                    title,
                    workbook.add_format(
                        {
                            "bold": True,
                            "font_size": 14,
                            "align": "center",
                            "font_name": font_name,
                        }
                    ),
                )
                row += 2  # Add spacing

            worksheet.merge_range(row, 0, row, 7, s_name, name_fmt)  # Col H is index 7
            row += 1
            # Ensure address is written to A4
            # Using simple write instead of merge if address is short, but merge is safer for multi-line
            if s_addr:
                set_val("A4", s_addr, align="left")
            else:
                # Fallback/Debug - No longer using hardcoded Bhopal address
                set_val("A4", "", align="left")

            set_val("A6", f"GSTIN/UIN: {s_gst}", bold_detail)
            set_val("A7", f"State Name : {s_state}, Code : {s_state_code}", bold_detail)

            # Explicitly force contact row if needed
            if s_phone:
                set_val("A5", f"Contact: {s_phone}", detail_fmt)
            row += 1
            worksheet.merge_range(row, 0, row, 7, f"Contact : {s_phone}", bold_detail)
            row += 1

        elif layout == "challan":
            # Layout matching 'DC12.xls'
            # Row 1: Tel (Left), GSTIN (Right)
            worksheet.write(row, 0, f"Tel. No. {s_phone}", tel_fmt)
            worksheet.merge_range(row, columns - 1, row, columns - 1, f"GSTIN: {s_gst}", tel_fmt)
            row += 2

            # Branding (Rows 3, 4, 5)
            worksheet.merge_range(row, 0, row, columns - 1, s_name, title_fmt)
            row += 1
            worksheet.merge_range(row, 0, row, columns - 1, s_desc, subtitle_fmt)
            row += 1
            worksheet.merge_range(row, 0, row, columns - 1, s_addr, subtitle_fmt)
            row += 2

            if title:
                worksheet.merge_range(
                    row,
                    0,
                    row,
                    columns - 1,
                    title,
                    workbook.add_format(
                        {
                            "bold": True,
                            "font_size": 14,
                            "align": "center",
                            "font_name": font_name,
                        }
                    ),
                )
                row += 1

        return row

    @staticmethod
    def _write_buyer_block(
        worksheet,
        workbook,
        row: int,
        col: int,
        db: sqlite3.Connection,
        header: dict = None,
        width: int = 5,
        label: str = "Buyer :",
        font_name: str = "Calibri",
    ):
        """
        Consistently writes the Buyer/Consignee block.
        Fetches from DB settings as default, overriden by specific record header if available.
        """
        # Fetch Supplier Settings
        try:
            rows = db.execute("SELECT key, value FROM settings").fetchall()
            settings = {row["key"]: row["value"] for row in rows}
        except Exception as e:
            logger.error(f"Failed to fetch business settings for buyer block: {e}")
            settings = {}

        # Fetch Default Buyer if not provided in header
        default_buyer = {}
        if not header.get("consignee_name"):
            try:
                buyer_row = db.execute("SELECT name, address, gstin, place_of_supply FROM buyers WHERE is_default = 1 LIMIT 1").fetchone()
                if buyer_row:
                    default_buyer = dict(buyer_row)
            except Exception as e:
                logger.error(f"Failed to fetch default buyer: {e}")

        # Default Buyer Info logic: Header > Default Buyer > Settings > Empty
        b_name = header.get("consignee_name") or default_buyer.get("name") or settings.get("buyer_name", "")
        b_addr = header.get("consignee_address") or default_buyer.get("address") or settings.get("buyer_address", "")
        b_gst = header.get("consignee_gstin") or default_buyer.get("gstin") or settings.get("buyer_gstin", "")

        # Parse Place of Supply if needed
        b_pos_raw = header.get("place_of_supply") or default_buyer.get("place_of_supply") or "BHOPAL, MP"
        b_pos = b_pos_raw

        # State logic (simple extraction if not provided)
        b_state = header.get("buyer_state") or "MP"
        if not header.get("buyer_state") and default_buyer.get("place_of_supply"):
            # Simple heuristic: last word or known states. For now defaulting to MP or extracting from POS
            pass

        # Formats - ALL buyer details should be BOLD with borders
        bold_border_fmt = workbook.add_format(
            {
                "bold": True,
                "font_size": 11,
                "font_name": font_name,
                "border": 1,
                "valign": "vcenter",
            }
        )

        if label:
            worksheet.merge_range(row, col, row, col + width, label, bold_border_fmt)
            row += 1

        # Each line should be in its own row with borders - NO merging
        worksheet.merge_range(row, col, row, col + width, b_name, bold_border_fmt)
        row += 1

        # Address might be multi-line but still one row with border
        worksheet.merge_range(row, col, row, col + width, b_addr, bold_border_fmt)
        row += 1

        # Empty row for spacing (as per template)
        worksheet.merge_range(row, col, row, col + width, "", bold_border_fmt)
        row += 1

        worksheet.merge_range(row, col, row, col + width, f"GSTIN/UIN : {b_gst}", bold_border_fmt)
        row += 1

        # Buyer state: only State Name, NO Code (as per template)
        worksheet.merge_range(row, col, row, col + width, f"State Name : {b_state}", bold_border_fmt)
        row += 1

        worksheet.merge_range(row, col, row, col + width, f"Place of Supply : {b_pos}", bold_border_fmt)
        row += 1

        return row + 1

    @staticmethod
    def generate_exact_invoice_excel(header: dict, items: list[dict], db: sqlite3.Connection, save_path: str = None):
        """
        Generate Invoice using 'GST_INV_31.xlsx' as a template.
        """
        # Template path: EXE-aware resolution
        import sys

        import openpyxl
        from openpyxl.styles import Alignment, Font
        from openpyxl.worksheet.cell_range import CellRange

        if getattr(sys, "frozen", False) and hasattr(sys, "_MEIPASS"):
            # Running inside PyInstaller EXE
            from pathlib import Path

            template_path = str(Path(sys._MEIPASS) / "backend" / "templates" / "GST_INV_31.xlsx")
        else:
            # Development mode
            template_path = os.path.join(os.path.dirname(__file__), "..", "templates", "GST_INV_31.xlsx")

        if not os.path.exists(template_path):
            logger.error("Template GST_INV_31.xlsx not found.")
            return StreamingResponse(io.BytesIO(b"Template not found"), media_type="text/plain")

        wb = openpyxl.load_workbook(template_path)
        ws = wb.active

        from openpyxl.cell.cell import MergedCell

        def set_val(coord, value, bold=False, align=None):
            try:
                cell = ws[coord]
                # If it's a merged cell, we must write to the master cell of the range
                if isinstance(cell, MergedCell):
                    for m_range in ws.merged_cells.ranges:
                        if coord in m_range:
                            cell = ws.cell(row=m_range.min_row, column=m_range.min_col)
                            break

                if value is not None:
                    cell.value = value
                if bold:
                    cell.font = Font(name="Calibri", size=10, bold=True)
                if align:
                    cell.alignment = Alignment(horizontal=align, vertical="vcenter", wrap_text=True)
            except Exception as e:
                logger.warning(f"Error setting value at {coord}: {e}")

        # 1. Fetch Settings
        try:
            rows = db.execute("SELECT key, value FROM settings").fetchall()
            settings = {row["key"]: row["value"] for row in rows}
        except Exception:
            settings = {}

        # 2. Header Info - Correct mappings based on template merge structure
        # Row 3: I3:K3=label, L3:N3=invoice#, O3=label "Dated", P3:T3=date value
        set_val("L3", header.get("invoice_number", ""), align="left")
        set_val("P3", header.get("invoice_date", ""), align="left")

        # Row 4: I4:N5=GEMC area, O4:P5="Date:" area, R4:T5=payment terms
        gem_num = header.get("gemc_number", "")
        if gem_num:
            set_val("I4", f"GEMC: {gem_num}", align="left")
            set_val("O4", f"Date: {header.get('gemc_date', '')}", align="left")
        set_val("R4", header.get("payment_terms", ""), align="left")

        # Row 6: I6:K7=label "Challan No", L6:N7=challan#, O6:O7=label, P6:T7=date
        set_val("L6", str(header.get("dc_number", "") or ""), align="left")
        set_val("P6", header.get("dc_date") or header.get("invoice_date", ""), align="left")

        # Row 8: I8:K9=label "Buyer's Order No", L8:N9=PO#, O8:O9=label, P8:T9=date
        set_val("L8", str(header.get("po_numbers", "") or ""), align="left")
        set_val("P8", str(header.get("buyers_order_date", "") or ""), align="left")

        # Row 10: I10:L11=label "Despatch Doc No", M10:N11=SRV area, O10=SRV#, Q10=SRV Dt
        set_val("L10", header.get("despatch_doc_no", ""), align="left")
        set_val("O10", header.get("srv_number") or header.get("srv_no", ""), align="left")
        srv_dt = header.get("srv_date", "")
        if srv_dt and len(srv_dt) > 10:
            srv_dt = srv_dt[:10]
        set_val("Q10", srv_dt, align="left")

        # Row 12: I12="Despatched through" label, L12=method value, O12="Destination" label, Q12=destination value
        set_val("L12", (header.get("despatch_through", "") or "BY LOADING VEHICLE").upper(), align="left")
        set_val("Q12", (header.get("destination", "") or "").upper(), align="left")  # Template has "Destination" label at O12

        # 3. Seller Info (Rows 3-7) - Template has labels, write values accordingly
        # Hardcoded fallbacks for Placeholder
        s_name = header.get("supplier_name") or settings.get("supplier_name", "") or "YOUR COMPANY NAME"
        s_addr = header.get("supplier_address") or settings.get("supplier_address", "") or "123, Industrial Area, Your City - 000000"
        s_gstin = header.get("supplier_gstin") or settings.get("supplier_gstin", "") or "23AAAAAAAAAAAAA"
        s_state = settings.get("supplier_state", "") or "State Name"
        s_code = settings.get("supplier_state_code", "") or "00"
        s_contact = header.get("supplier_contact") or settings.get("supplier_contact", "") or "+91 00000 00000"

        # Row 3: Supplier Name (A3:F3 merged placeholder)
        set_val("A3", s_name, align="left")
        # Row 4: Supplier Address (A4:F4 merged placeholder)
        set_val("A4", s_addr, align="left")
        # Row 5: GSTIN value (A5:B5=label merged, C5:F5=value merged)
        set_val("C5", s_gstin, align="left")
        # Row 6: State value (A6:B6=label merged, C6:D6=value merged), Code (E6=label, F6=value)
        set_val("C6", s_state, align="left")
        set_val("F6", s_code, align="left")
        # Row 7: Contact value (A7:B7=label merged, C7 onwards=value)
        set_val("C7", s_contact, align="left")

        # 4. Buyer Info (Rows 8-13) - Template has labels, write values accordingly
        # Hardcoded fallbacks for default buyer
        b_name = header.get("consignee_name") or header.get("buyer_name")
        b_addr = header.get("consignee_address") or header.get("buyer_address")
        b_gst = header.get("consignee_gstin") or header.get("buyer_gstin")
        b_place = header.get("place_of_supply") or "BHOPAL, MP"
        b_state = header.get("buyer_state") or "MP"

        # FALLBACK: If Buyer Name is missing in header, fetch Default Buyer from DB
        if not b_name:
            try:
                buyer_row = db.execute("SELECT name, address, gstin, place_of_supply, state FROM buyers WHERE is_default = 1 LIMIT 1").fetchone()
                if buyer_row:
                    b_name = buyer_row["name"]
                    b_addr = buyer_row["address"]
                    b_gst = buyer_row["gstin"]
                    b_place = buyer_row["place_of_supply"]
                    b_state = buyer_row["state"]
            except Exception as e:
                logger.error(f"Failed to fetch default buyer in Excel: {e}")

        # Final hardcoded fallbacks if still missing
        b_name = b_name or "M/S Partner Engineering PSU Ltd."
        b_addr = b_addr or "Sr. Accounts Officer (Purchase)"
        b_gst = b_gst or "23BBBBBBBBBBBBB"
        b_state = b_state or "State"
        b_place = b_place or "CITY, STATE"

        # Row 8: "Buyer" label already in template - keep as-is
        # Row 9: Buyer Name (A9:F9 merged placeholder)
        set_val("A9", b_name, align="left")
        # Row 10: Buyer Address (A10:F10 merged placeholder)
        set_val("A10", b_addr, align="left")
        # Row 11: GSTIN value (A11:B11=label merged, C11:F11=value merged)
        set_val("C11", b_gst, align="left")
        # Row 12: State value (A12:B12=label merged, C12:F12=value merged)
        set_val("C12", b_state, align="left")
        # Row 13: Place of Supply value (A13:B13=label merged, C13:F13=value merged)
        set_val("C13", b_place, align="left")

        # 5. Line Items (Start at Row 17)
        start_row = 17
        template_total_row = 23  # Fixed template position
        template_capacity = template_total_row - start_row  # 6
        num_items = len(items)

        # MAT CODE column is now pre-built in template at Column J
        # No dynamic insertion needed - template GST_INV_31.xlsx already has the column

        # Calculate expansion only (No negative adjustment for deletion)
        rows_to_insert = max(0, num_items - template_capacity)

        # 1. Clear the pre-allocated area content
        for r_clear in range(start_row, template_total_row):
            for c_clear in range(1, 20):
                cell = ws.cell(row=r_clear, column=c_clear)
                if not isinstance(cell, MergedCell):
                    cell.value = None

        # 2. Expand if needed
        if rows_to_insert > 0:
            try:
                # Capture merges
                m_ranges = [CellRange(m.coord) for m in ws.merged_cells.ranges]
                ws.merged_cells.clear()

                ws.insert_rows(start_row + 1, amount=rows_to_insert)

                # Restore/Shift merges
                for m in m_ranges:
                    if m.min_row >= start_row + 1:
                        m.shift(row_shift=rows_to_insert)
                    ws.merged_cells.add(m)
            except Exception as e:
                logger.error(f"Failed to insert rows: {e}")

        # 3. Shrink (Hide Rows) if items < capacity
        # This replaces "Delete" logic which was unstable.
        # If items=3. We fill 17,18,19. We hide 20,21,22.
        elif num_items < template_capacity:
            # Range to hide: from (start + num_items) up to (template_total_row)
            # e.g. 3 items: start(17)+3 = 20. hide 20, 21, 22.
            hide_start = start_row + num_items
            for r_hide in range(hide_start, template_total_row):
                ws.row_dimensions[r_hide].hidden = True

        # 4. Write Items
        t_qty = 0
        t_taxable = 0
        t_cgst = 0
        t_sgst = 0
        t_total = 0

        cgst_rate_val = float(settings.get("cgst_rate", 9.0))
        sgst_rate_val = float(settings.get("sgst_rate", 9.0))

        for idx, item in enumerate(items):
            r = start_row + idx

            # For new rows (if we expanded), we need style copy
            # For new rows (if we expanded), we need style copy
            if r > start_row and rows_to_insert > 0:
                # Logic for copying style only if we inserted rows
                # Extended range to 26 (Z) to account for extra column
                for col in range(1, 26):
                    src_c = ws.cell(row=start_row, column=col)
                    dest_c = ws.cell(row=r, column=col)
                    if src_c.has_style:
                        dest_c.font = Font(name=src_c.font.name, size=src_c.font.size, bold=src_c.font.bold)
                        # Only copy border if not None?
                        dest_c.border = copy(src_c.border)
                        dest_c.alignment = copy(src_c.alignment)
                try:
                    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8)
                except Exception:
                    pass

            qty = float(item.get("quantity", 0) or 0)
            rate = float(item.get("rate", 0) or 0)
            taxable = float(item.get("taxable_value", qty * rate))
            cgst = float(item.get("cgst_amount", 0))
            sgst = float(item.get("sgst_amount", 0))
            line_total = float(item.get("total_amount", taxable + cgst + sgst))

            t_qty += qty
            t_taxable += taxable
            t_cgst += cgst
            t_sgst += sgst
            t_total += line_total

            description = str(item.get("description") or "No Description")

            # Mat Code - already in separate column, no need to strip from description
            mat_code = item.get("material_code") or ""

            set_val(f"A{r}", idx + 1, align="center")
            set_val(f"B{r}", description, align="left")
            set_val(f"I{r}", item.get("hsn_sac") or "")

            # MAT CODE: Dedicated Column J
            set_val(f"J{r}", mat_code, align="center")

            # SHIFTED INDICES DUE TO NEW COLUMN J
            set_val(f"K{r}", 1, align="center")  # No of Pckt
            set_val(f"L{r}", qty, align="center")
            set_val(f"M{r}", rate, align="center")
            set_val(f"N{r}", item.get("unit", "NOS"), align="center")
            set_val(f"O{r}", taxable, align="right")
            set_val(f"P{r}", cgst_rate_val, align="center")
            set_val(f"Q{r}", cgst, align="right")
            set_val(f"R{r}", sgst_rate_val, align="center")
            set_val(f"S{r}", sgst, align="right")
            set_val(f"T{r}", line_total, align="right")

        # 4. Remove Logic for "Clear empty pre-allocated rows"
        # (It's handled by delete_rows above)

        # 6. Totals Area
        # Determine actual footer start positions
        # If hiding rows, insert is 0, so total stays at 23 (template_total_row).
        # If expanding, insert is >0, so total shifts down.
        final_total_row = template_total_row + rows_to_insert

        # User requested Total Amount at Row 23 (if no insert)
        # Explicitly write "Total" label if it was cleared
        set_val(f"H{final_total_row}", "Total", bold=True, align="right")

        # Write Numerical Totals
        # Write Numerical Totals (Shifted for Column J)
        set_val(f"L{final_total_row}", t_qty)
        set_val(f"O{final_total_row}", t_taxable)
        set_val(f"Q{final_total_row}", t_cgst)
        set_val(f"S{final_total_row}", t_sgst)
        set_val(f"T{final_total_row}", t_total)

        # Words row (Below Total)
        words_row = final_total_row + 1
        # Fix: Remove "Rupees" prefix as amount_to_words includes it
        set_val(f"A{words_row}", f"Total Amount (In Words):- {amount_to_words(t_total)}", align="left")

        # Tax Summary Table
        # Base (Total)
        # Base + 1: Words
        # Base + 2: Tax Headers (Preserve)
        # Base + 3: Tax Values (Title: Taxable Value | Rate | Amount)

        tax_sum_row = final_total_row + 3
        set_val(f"O{tax_sum_row}", t_taxable)
        set_val(f"P{tax_sum_row}", cgst_rate_val)
        set_val(f"Q{tax_sum_row}", t_cgst)
        set_val(f"R{tax_sum_row}", sgst_rate_val)
        set_val(f"S{tax_sum_row}", t_sgst)
        set_val(f"T{tax_sum_row}", t_cgst + t_sgst)

        # Tax Words
        # Base + 5: SGST Words
        # Base + 6: CGST Words

        tax_words_row = final_total_row + 5

        # Write clean Tax Words
        set_val(f"A{tax_words_row}", f"SGST (in words) : {amount_to_words(t_sgst)}", align="left")
        set_val(f"A{tax_words_row + 1}", f"CGST (in words) : {amount_to_words(t_cgst)}", align="left")

        # 7. Declaration & Signatory
        # Template already has signatory placeholders pre-printed
        # No need to add them again - they would duplicate

        output = io.BytesIO()
        wb.save(output)

        filename = f"Invoice_{header.get('invoice_number')}.xlsx"
        return ExcelService._save_or_stream(output, filename, save_path)

    @staticmethod
    def generate_dispatch_summary(date_str: str, items: list[dict], db: sqlite3.Connection, save_path: str = None):
        """
        Generate strict Excel format matching 'Summary.xls' and User Screenshot
        """
        output = io.BytesIO()
        workbook = xlsxwriter.Workbook(output)
        worksheet = workbook.add_worksheet("Summary")

        # Styles
        workbook.add_format({"bold": True, "font_size": 18, "align": "center", "font_name": "Calibri"})
        workbook.add_format({"bold": True, "font_size": 11, "align": "center", "font_name": "Calibri"})
        header_table = workbook.add_format(
            {
                "bold": True,
                "border": 1,
                "align": "center",
                "valign": "vcenter",
                "text_wrap": True,
                "font_name": "Calibri",
            }
        )
        cell_fmt = workbook.add_format(
            {
                "border": 1,
                "align": "center",
                "valign": "vcenter",
                "font_name": "Calibri",
            }
        )
        bold_left = workbook.add_format({"bold": True, "font_name": "Calibri"})

        # Column Widths
        worksheet.set_column("A:A", 5)  # S.No.
        worksheet.set_column("B:B", 30)  # Description
        worksheet.set_column("C:C", 15)  # Quantity
        worksheet.set_column("D:D", 8)  # No of packets
        worksheet.set_column("E:E", 12)  # PO NO
        worksheet.set_column("F:F", 18)  # GEMC NO
        worksheet.set_column("G:G", 10)  # Invoice No.
        worksheet.set_column("H:H", 10)  # Challan No.
        worksheet.set_column("I:I", 12)  # Dispatch Delivered

        # Header Section
        current_row = ExcelService._write_standard_header(worksheet, workbook, columns=9, db=db, title="SUMMARY", layout="challan")

        worksheet.write(current_row, 0, "Date:", bold_left)
        worksheet.write(current_row, 1, date_str, bold_left)

        table_row = current_row + 2
        headers = [
            "S. No.",
            "Description",
            "Quantity Set/Nos.",
            "No of packets",
            "PO NO",
            "GEMC NO",
            "Invoice No.",
            "Challan No.",
            "Dispatch Delivered",
        ]
        for i, h in enumerate(headers):
            worksheet.write(table_row, i, h, header_table)

        # Data
        row = table_row + 1
        for idx, item in enumerate(items):
            worksheet.write(row, 0, idx + 1, cell_fmt)
            worksheet.write(row, 1, item.get("description", ""), cell_fmt)
            worksheet.write(row, 2, f"{item.get('quantity', '')} {item.get('unit', '')}", cell_fmt)
            worksheet.write(row, 3, item.get("no_of_packets", ""), cell_fmt)
            worksheet.write(row, 4, item.get("po_number", ""), cell_fmt)
            worksheet.write(row, 5, item.get("gemc_number", ""), cell_fmt)
            worksheet.write(row, 6, item.get("invoice_number", ""), cell_fmt)
            worksheet.write(row, 7, item.get("dc_number", ""), cell_fmt)
            worksheet.write(row, 8, item.get("destination", ""), cell_fmt)
            row += 1

        workbook.close()
        filename = f"Summary_{date_str}.xlsx"
        return ExcelService._save_or_stream(output, filename, save_path)

    @staticmethod
    def generate_technical_summary(items: list[dict], db: sqlite3.Connection, save_path: str = None) -> StreamingResponse:
        """
        Generate Excel export matching the 'TECHNICAL SUMMARY' template.
        Columns: S.No, Description, Qty Set/Nos, Packets, PO NO, GEMC NO, Invoice No, Challan No, Dispatch Delivered.
        """
        output = io.BytesIO()
        settings = {}  # Fixed: Defined variable to support get() calls later
        workbook = xlsxwriter.Workbook(output)
        worksheet = workbook.add_worksheet("Summary")

        # --- Formats ---
        font_name = "Calibri"

        # Header: YOUR COMPANY (Bold, 14pt-16pt)
        fmt_company = workbook.add_format({"bold": True, "font_size": 16, "align": "center", "font_name": font_name})
        # SubHeader: Address details (Bold, 10pt)
        fmt_address = workbook.add_format({"bold": True, "font_size": 10, "align": "center", "font_name": font_name, "text_wrap": True})
        # Title: SUMMARY (Bold, 14pt)
        fmt_title = workbook.add_format({"bold": True, "font_size": 14, "align": "center", "font_name": font_name, "top": 1, "bottom": 1})

        # Date Row
        fmt_date_label = workbook.add_format({"bold": True, "font_name": font_name, "align": "right"})
        fmt_date_val = workbook.add_format({"bold": True, "font_name": font_name, "align": "left"})

        # Table Header
        fmt_th = workbook.add_format({"bold": True, "border": 1, "align": "center", "valign": "vcenter", "text_wrap": True, "font_name": font_name})

        # Table Cells
        fmt_cell = workbook.add_format({"border": 1, "align": "center", "valign": "vcenter", "font_name": font_name, "text_wrap": True})
        fmt_cell_left = workbook.add_format({"border": 1, "align": "left", "valign": "vcenter", "font_name": font_name, "text_wrap": True})

        # Footer
        fmt_footer = workbook.add_format({"bold": True, "font_name": font_name, "align": "right"})

        # --- Column Widths ---
        # A=S.No(5), B=Desc(40), C=Qty(10), D=Pkts(8), E=PO(12), F=GEMC(15), G=Inv(10), H=Chal(10), I=Disp(10)
        worksheet.set_column("A:A", 5)
        worksheet.set_column("B:B", 40)
        worksheet.set_column("C:C", 12)
        worksheet.set_column("D:D", 8)
        worksheet.set_column("E:E", 14)
        worksheet.set_column("F:F", 18)
        worksheet.set_column("G:G", 12)
        worksheet.set_column("H:H", 12)
        worksheet.set_column("I:I", 12)

        # --- Header Section ---
        # 0. Fetch Settings
        try:
            rows = db.execute("SELECT key, value FROM settings").fetchall()
            settings = {row["key"]: row["value"] for row in rows}
        except Exception:
            settings = {}

        # Row 0: Tel (Left)
        s_phone = settings.get("supplier_contact") or settings.get("supplier_phone", "")
        worksheet.write(0, 0, f"Tel. No. {s_phone}", workbook.add_format({"font_name": font_name}))

        # Row 2: COMPANY NAME from Settings
        comp_name = settings.get("supplier_name", settings.get("company_name", ""))
        worksheet.merge_range("A3:I3", comp_name, fmt_company)

        # Row 3: Description
        s_desc = settings.get("supplier_description", "")
        worksheet.merge_range("A4:I4", s_desc, fmt_address)

        # Row 4: Address
        s_addr = settings.get("supplier_address", "")
        worksheet.merge_range("A5:I5", s_addr, fmt_address)

        # Row 5: SUMMARY
        worksheet.merge_range("A6:I6", "SUMMARY", fmt_title)

        # Row 6: Date
        from datetime import datetime

        current_date = datetime.now().strftime("%d-%m-%Y")
        worksheet.write(6, 1, "Date:", fmt_date_label)  # B7
        worksheet.write(6, 2, current_date, fmt_date_val)  # C7

        # --- Table Headers (Row 8 -> Index 7) ---
        headers = [
            "S.\nNo.",
            "Description",
            "Quantity\nSet/Nos.",
            "No of\npackets",
            "PO NO",
            "GEMC  NO",
            "Invoice\nNo.",
            "Challa\nn\nNo.",
            "Dispatch\nDelivered",
        ]

        start_row = 8
        for i, h in enumerate(headers):
            worksheet.write(start_row, i, h, fmt_th)

        # --- Data Loop ---
        row = start_row + 1
        total_qty = 0  # This might be mixed units, but let's try
        total_pkts = 0

        for idx, item in enumerate(items):
            # Parse qty for total if possible
            q_val = item.get("ord_qty", 0)
            try:
                # Handle potential string/float inputs safely
                if isinstance(q_val, str):
                    q_val = float(q_val.replace(",", ""))
                total_qty += int(q_val)
            except (ValueError, TypeError):
                pass

            p_val = item.get("no_of_packets", 0)
            try:
                if isinstance(p_val, str):
                    p_val = float(p_val.replace(",", ""))
                total_pkts += int(p_val)
            except (ValueError, TypeError):
                pass

            unit = item.get("unit", "")
            qty_display = f"{q_val} {unit}".strip()

            worksheet.write(row, 0, idx + 1, fmt_cell)
            worksheet.write(row, 1, item.get("description", ""), fmt_cell_left)  # Align Left for Desc
            worksheet.write(row, 2, qty_display, fmt_cell)
            worksheet.write(row, 3, p_val if p_val else "", fmt_cell)
            worksheet.write(row, 4, item.get("po_number", ""), fmt_cell)
            worksheet.write(row, 5, item.get("gemc_number", ""), fmt_cell)
            worksheet.write(row, 6, item.get("invoice_number", ""), fmt_cell)
            worksheet.write(row, 7, item.get("dc_number", ""), fmt_cell)
            worksheet.write(row, 8, item.get("dispatch_delivered", ""), fmt_cell)
            row += 1

        # --- Footer ---
        # Total Row
        worksheet.write(row, 1, "Total", fmt_th)  # Or just format
        # Logic for total mix units: "3440 Nos + 1 Set" if logical, complex.
        # For now, just leave blank or basic sum if uniform.
        # User image shows "3440 Nos + 1 Set".
        # We will iterate and build a summary string if multiple units.

        # Build unit summary
        unit_sums = {}
        for item in items:
            u = item.get("unit", "Nos")
            try:
                val = float(item.get("ord_qty", 0))
                unit_sums[u] = unit_sums.get(u, 0) + val
            except Exception:
                pass

        summary_str = " + ".join([f"{v:g} {k}" for k, v in unit_sums.items()])

        worksheet.write(row, 2, summary_str, fmt_th)  # Quantity col
        worksheet.write(row, 3, total_pkts, fmt_th)  # Packets col

        # Apply borders to empty cells in total row
        for i in [0, 4, 5, 6, 7, 8]:
            worksheet.write(row, i, "", fmt_th)

        # Signature
        row += 2
        comp_disp = settings.get("supplier_name", settings.get("company_name", ""))
        worksheet.merge_range(row, 5, row, 8, f"For {comp_disp}", fmt_footer)

        workbook.close()
        output.seek(0)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"Summary_{timestamp}.xlsx"

        if save_path:
            return ExcelService._save_or_stream(output, filename, save_path)

        headers = {"Content-Disposition": f'attachment; filename="{filename}"'}
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers=headers,
        )

    @staticmethod
    def generate_from_list(data: list[dict], filename: str, save_path: str = None) -> StreamingResponse:
        """
        Generic helper to stream list of dicts as Excel or save to disk.
        Replaces generate_from_df.
        """
        import io

        import xlsxwriter

        output = io.BytesIO()
        workbook = xlsxwriter.Workbook(output, {"in_memory": True})
        worksheet = workbook.add_worksheet("Report")

        if not data:
            workbook.close()
            output.seek(0)
            return ExcelService._save_or_stream(output, filename, save_path)

        # Get headers from first item keys
        headers = list(data[0].keys())

        # Write Headers
        header_fmt = workbook.add_format({"bold": True, "border": 1})
        for col_num, header in enumerate(headers):
            worksheet.write(0, col_num, header, header_fmt)

        # Write Data
        for row_num, item in enumerate(data, start=1):
            for col_num, header in enumerate(headers):
                value = item.get(header, "")
                worksheet.write(row_num, col_num, value)

        # Auto-adjust columns (rough approximation)
        for i, header in enumerate(headers):
            max_len = len(str(header))
            for item in data:
                val_len = len(str(item.get(header, "")))
                if val_len > max_len:
                    max_len = val_len
            worksheet.set_column(i, i, min(max_len + 2, 50))  # Cap at 50

        workbook.close()
        output.seek(0)
        return ExcelService._save_or_stream(output, filename, save_path)

    @staticmethod
    def generate_pending_items_report(items: list[dict], filename: str, db: sqlite3.Connection, save_path: str = None):
        # Use generic list export
        return ExcelService.generate_from_list(items, filename, save_path)

    @staticmethod
    def generate_po_upload_template() -> StreamingResponse:
        """
        Generate empty PO upload template with required headers
        """
        import io

        import xlsxwriter

        output = io.BytesIO()
        workbook = xlsxwriter.Workbook(output, {"in_memory": True})
        worksheet = workbook.add_worksheet("PO_Upload")

        headers = [
            "PO Number",
            "PO Date",
            "Vendor Name",
            "Project Name",
            "Item No",
            "Material Code",
            "Description",
            "Unit",
            "Qty",
            "Rate",
            "Delivery Date",
        ]

        header_fmt = workbook.add_format({"bold": True, "bg_color": "#D7E4BC", "border": 1})
        for col_num, value in enumerate(headers):
            worksheet.write(0, col_num, value, header_fmt)

        worksheet.set_column("A:K", 15)

        workbook.close()
        output.seek(0)

        return StreamingResponse(
            output,
            headers={"Content-Disposition": 'attachment; filename="PO_Upload_Template.xlsx"'},
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    @staticmethod
    def generate_srv_upload_template() -> StreamingResponse:
        """
        Generate empty SRV upload template with required headers
        """
        import io

        import xlsxwriter

        output = io.BytesIO()
        workbook = xlsxwriter.Workbook(output, {"in_memory": True})
        worksheet = workbook.add_worksheet("SRV_Upload")

        headers = [
            "SRV Number",
            "SRV Date",
            "PO Number",
            "PO Item No",
            "Lot No",
            "Received Qty",
            "Rejected Qty",
            "Challan No",
            "Challan Date",
            "Invoice No",
            "Invoice Date",
            "Remarks",
        ]

        header_fmt = workbook.add_format({"bold": True, "bg_color": "#DDEBF7", "border": 1})
        for col_num, value in enumerate(headers):
            worksheet.write(0, col_num, value, header_fmt)

        worksheet.set_column("A:L", 15)

        workbook.close()
        output.seek(0)

        return StreamingResponse(
            output,
            headers={"Content-Disposition": 'attachment; filename="SRV_Upload_Template.xlsx"'},
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    @staticmethod
    def generate_exact_dc_excel(header: dict, items: list[dict], db: sqlite3.Connection, save_path: str = None):
        """
        Generate Delivery Challan matching the optimized single-page A4 printout format.
        Uses xlsxwriter for total control over layout.
        """
        import io
        from datetime import datetime

        import xlsxwriter

        output = io.BytesIO()
        workbook = xlsxwriter.Workbook(output, {"in_memory": True})
        worksheet = workbook.add_worksheet("Delivery Challan")

        # --- Print Settings for A4 Single Page ---
        worksheet.set_paper(9)  # A4
        worksheet.fit_to_pages(1, 1)  # Fit to 1 page
        worksheet.set_margins(left=0.5, right=0.5, top=0.3, bottom=0.3)
        # Portrait Mode

        # --- Formats ---
        base_font = "Calibri"

        fmt_tel = workbook.add_format({"font_size": 11, "font_name": base_font, "valign": "vcenter"})
        fmt_gstin = workbook.add_format({"font_size": 11, "font_name": base_font, "valign": "vcenter", "align": "right"})
        fmt_title = workbook.add_format({"bold": True, "font_size": 22, "align": "center", "font_name": base_font})
        fmt_branding = workbook.add_format({"bold": True, "font_size": 11, "align": "center", "font_name": base_font})
        fmt_address = workbook.add_format({"font_size": 10, "align": "center", "font_name": base_font})
        fmt_doc_type = workbook.add_format({"bold": True, "font_size": 16, "align": "center", "border": 1, "font_name": base_font})

        fmt_label = workbook.add_format({"font_size": 10, "font_name": base_font, "valign": "vcenter"})
        fmt_value = workbook.add_format({"font_size": 10, "bold": True, "font_name": base_font, "valign": "vcenter"})

        fmt_box_l = workbook.add_format({"font_size": 10, "font_name": base_font, "valign": "vcenter", "left": 1})
        fmt_box_r = workbook.add_format({"font_size": 10, "font_name": base_font, "valign": "vcenter", "right": 1})
        fmt_box_tl = workbook.add_format({"font_size": 10, "font_name": base_font, "valign": "vcenter", "left": 1, "top": 1})
        fmt_box_tr = workbook.add_format({"font_size": 10, "font_name": base_font, "valign": "vcenter", "right": 1, "top": 1})
        fmt_box_bl = workbook.add_format({"font_size": 10, "font_name": base_font, "valign": "vcenter", "left": 1, "bottom": 1})
        fmt_box_br = workbook.add_format({"font_size": 10, "font_name": base_font, "valign": "vcenter", "right": 1, "bottom": 1})
        fmt_box_t = workbook.add_format({"font_size": 10, "font_name": base_font, "valign": "vcenter", "top": 1})
        fmt_box_b = workbook.add_format({"font_size": 10, "font_name": base_font, "valign": "vcenter", "bottom": 1})

        # Border fixes for middle values in Right Box
        fmt_val_t = workbook.add_format({"font_size": 10, "bold": True, "font_name": base_font, "valign": "vcenter", "top": 1})
        fmt_val_b = workbook.add_format({"font_size": 10, "bold": True, "font_name": base_font, "valign": "vcenter", "bottom": 1})

        fmt_th = workbook.add_format(
            {"bold": True, "border": 1, "align": "center", "valign": "vcenter", "font_name": base_font, "text_wrap": True, "font_size": 11}
        )
        fmt_td = workbook.add_format({"border": 1, "valign": "top", "font_name": base_font, "text_wrap": True, "font_size": 10})
        fmt_td_center = workbook.add_format({"border": 1, "align": "center", "valign": "vcenter", "font_name": base_font, "font_size": 10})
        # Changed to border: 1 to ensure horizontal lines appear between remarks
        fmt_td_remark = workbook.add_format({"border": 1, "valign": "top", "font_name": base_font, "font_size": 10})
        fmt_footer = workbook.add_format({"font_size": 10, "font_name": base_font, "valign": "vcenter"})
        fmt_sig = workbook.add_format({"bold": True, "font_size": 11, "font_name": base_font, "align": "right"})

        # Column Widths (A-H for 8 columns)
        worksheet.set_column("A:A", 12)  # P.O.SI. No.
        worksheet.set_column("B:F", 12)  # Description spans B-F
        worksheet.set_column("G:G", 15)  # Qty / Labels
        worksheet.set_column("H:H", 18)  # Values

        # --- Helper Functions ---
        def fmt_date(d_str):
            if not d_str:
                return ""
            try:
                dt = datetime.strptime(str(d_str).split("T")[0], "%Y-%m-%d")
                return dt.strftime("%d/%m/%Y")
            except:
                return str(d_str)

        # --- Fetch Settings ---
        try:
            rows = db.execute("SELECT key, value FROM settings").fetchall()
            settings = {row["key"]: row["value"] for row in rows}
        except Exception:
            settings = {}

        s_phone = settings.get("supplier_contact") or settings.get("supplier_phone", "")
        s_gst = settings.get("supplier_gstin", "")
        s_name = settings.get("supplier_name", "YOUR COMPANY")
        s_addr = settings.get("supplier_address", "")

        # --- Row 1: Tel and GSTIN ---
        worksheet.write("A1", f"Tel. No. {s_phone}", fmt_tel)
        worksheet.write("H1", f"GSTIN: {s_gst}", fmt_gstin)

        # --- Row 3: Company Name (Centered) ---
        worksheet.merge_range("A3:H3", s_name, fmt_title)

        # --- Row 4: HARDCODED Branding Line ---
        worksheet.merge_range("A4:H4", "Specialist Manufacturers and Suppliers of Engineering Solutions", fmt_branding)

        # --- Row 5: Address ---
        worksheet.merge_range("A5:H5", s_addr, fmt_address)

        # --- Row 7: DELIVERY CHALLAN Title ---
        worksheet.merge_range("A7:H7", "DELIVERY CHALLAN", fmt_doc_type)

        # --- Rows 8-10: Info Boxes ---
        # Left Box: Consignee (A-E)
        b_name = header.get("consignee_name") or "The Purchase Manager"
        b_company = "M/S Partner Engineering PSU Ltd."
        b_location = "CITY NAME"

        worksheet.write("A8", b_name, fmt_box_tl)
        worksheet.write("E8", "", fmt_box_tr)
        worksheet.write("A9", b_company, fmt_box_l)
        worksheet.write("E9", "", fmt_box_r)
        worksheet.write("A10", b_location, fmt_box_bl)
        worksheet.write("E10", "", fmt_box_br)
        for c in range(1, 5):
            worksheet.write(7, c, "", fmt_box_t)
            worksheet.write(9, c, "", fmt_box_b)

        # Right Box: Challan Info (F-H)
        worksheet.write("F8", "Challan No. :", fmt_box_tl)
        worksheet.write("G8", header.get("dc_number", ""), fmt_val_t)  # Top border
        worksheet.write("H8", "", fmt_box_tr)

        worksheet.write("F9", "Date :", fmt_box_l)
        worksheet.write("G9", fmt_date(header.get("dc_date")), fmt_value)
        worksheet.write("H9", "", fmt_box_r)

        worksheet.write("F10", "Our Ref :", fmt_box_bl)
        worksheet.write("G10", header.get("our_ref", ""), fmt_val_b)  # Bottom border (Fix 1)
        worksheet.write("H10", "", fmt_box_br)

        # --- Row 12: Order Info (No borders as requested) ---
        worksheet.write("A12", f"Your Order No. :  {header.get('po_number', '')}", fmt_label)
        worksheet.write("D12", "Date:", fmt_label)
        worksheet.write("E12", fmt_date(header.get("po_date")), fmt_value)
        worksheet.write("F12", "Amd. Date:", fmt_label)

        # --- Row 13: Goods Dispatched ---
        dest_code = header.get("department_no") or header.get("destination") or ""
        worksheet.write("A13", f"Goods Dispatched Delivered to:  {dest_code}", fmt_label)

        # --- Row 15: Table Header ---
        tr = 14  # 0-indexed row 14 = Excel row 15
        worksheet.write(tr, 0, "P.O.SI.\nNo.", fmt_th)
        worksheet.merge_range(tr, 1, tr, 5, "Description", fmt_th)
        worksheet.merge_range(tr, 6, tr, 7, "Quantity", fmt_th)
        worksheet.set_row(tr, 35)

        # --- Items ---
        curr = tr + 1
        total_qty = 0
        for idx, item in enumerate(items):
            qty_val = float(item.get("dsp_qty") or 0)
            total_qty += qty_val
            q_num = int(qty_val) if qty_val == int(qty_val) else qty_val
            qty_str = f"{q_num} {item.get('unit', 'NO')}"

            desc = item.get("material_description") or item.get("description", "")
            worksheet.set_row(curr, max(35, 15 * (1 + len(desc) // 60)))

            worksheet.write(curr, 0, item.get("po_item_no", idx + 1), fmt_td_center)
            worksheet.merge_range(curr, 1, curr, 5, desc, fmt_td)
            worksheet.merge_range(curr, 6, curr, 7, qty_str, fmt_td_center)
            curr += 1

        # --- Remarks/Notes Section ---
        remarks_list = []

        # Auto-inject GC number
        gc_no = header.get("gc_number")
        gc_dt = header.get("gc_date")
        if gc_no:
            remarks_list.append(f"Guarantee Certificate No. {gc_no} Dt. {fmt_date(gc_dt)}")

        # GST Bill No - Always include as per user request
        inv_no = header.get("invoice_number", "")
        inv_dt = header.get("invoice_date")
        # Ensure distinct spacing or format if needed, but standard text is fine
        gst_str = f"GST Bill No. {inv_no}  Dt. {fmt_date(inv_dt) if inv_dt else ''}"
        remarks_list.append(gst_str)

        # Standard template lines
        remarks_list.extend(["Dimension Report", "TC No:-  dt.  Of", "TC No  dt.  Of", "Lot No.  -"])

        # Add spacing for consignment value
        total_value = sum(float(i.get("dsp_qty", 0) or 0) * float(i.get("rate", 0) or i.get("po_rate", 0) or 0) for i in items)
        remarks_list.append(f"Consignment Value of PSU DC ₹{total_value:,.2f}")

        for rem_text in remarks_list:
            worksheet.set_row(curr, 18)
            worksheet.write(curr, 0, "", fmt_td_remark)
            worksheet.merge_range(curr, 1, curr, 5, rem_text, fmt_td_remark)
            worksheet.merge_range(curr, 6, curr, 7, "", fmt_td_remark)
            curr += 1

        # Closing border for table (Full width A-H) (Fix 2)
        for c in range(0, 8):
            worksheet.write(curr, c, "", workbook.add_format({"top": 1}))

        # --- Footer ---
        footer_r = curr + 1
        worksheet.write(footer_r, 1, "Received the Goods in good condition", fmt_footer)
        s_name = settings.get("supplier_name", settings.get("company_name", ""))
        worksheet.merge_range(footer_r, 6, footer_r, 7, f"For {s_name}", fmt_sig)

        worksheet.write(footer_r + 2, 1, "E. & O.E.", fmt_footer)
        # Authorised Signatory removed

        workbook.close()
        output.seek(0)

        filename = f"DC_{header.get('dc_number', 'Draft')}.xlsx"
        return ExcelService._save_or_stream(output, filename, save_path)

    @staticmethod
    def generate_gc_excel(header: dict, items: list[dict], db: sqlite3.Connection, save_path: str = None) -> StreamingResponse:
        """
        Generate Guarantee Certificate matchin the user provided screenshot.
        Portrait A4.
        """
        import io
        from datetime import datetime

        import xlsxwriter

        output = io.BytesIO()
        workbook = xlsxwriter.Workbook(output, {"in_memory": True})
        worksheet = workbook.add_worksheet("Guarantee Certificate")

        # --- Print Settings ---
        worksheet.set_paper(9)  # A4
        worksheet.fit_to_pages(1, 1)  # Fit to 1x1 page
        worksheet.set_margins(left=0.5, right=0.5, top=0.5, bottom=0.5)
        # Portrait (default)

        # --- Formats ---
        font_name = "Calibri"

        fmt_tel = workbook.add_format({"font_name": font_name, "font_size": 11, "valign": "vcenter"})
        fmt_company = workbook.add_format({"bold": True, "font_name": font_name, "font_size": 20, "align": "center"})
        fmt_branding = workbook.add_format({"bold": True, "font_name": font_name, "font_size": 11, "align": "center"})
        fmt_title = workbook.add_format({"bold": True, "font_name": font_name, "font_size": 16, "align": "center"})

        fmt_box = workbook.add_format({"border": 1, "font_name": font_name, "font_size": 11, "valign": "vcenter"})

        fmt_table_header = workbook.add_format(
            {"bold": True, "border": 1, "align": "center", "valign": "vcenter", "font_name": font_name, "font_size": 11, "text_wrap": True}
        )
        fmt_cell_center = workbook.add_format({"border": 1, "align": "center", "valign": "vcenter", "font_name": font_name, "font_size": 11})
        fmt_cell_desc = workbook.add_format(
            {"border": 1, "align": "left", "valign": "vcenter", "font_name": font_name, "font_size": 11, "text_wrap": True}
        )

        fmt_footer_text = workbook.add_format(
            {"border": 1, "valign": "vcenter", "text_wrap": True, "align": "left", "font_name": font_name, "font_size": 11}
        )

        # --- Column Widths ---
        # Based on image: A is narrow, B-E wide, F wide
        worksheet.set_column("A:A", 10)
        worksheet.set_column("B:E", 12)
        worksheet.set_column("F:F", 25)

        # --- Helper for dates ---
        def fmt_date(d_str):
            if not d_str:
                return ""
            try:
                dt = datetime.strptime(str(d_str).split("T")[0], "%Y-%m-%d")
                return dt.strftime("%d/%m/%Y")
            except:
                return str(d_str)

        # --- Header ---
        # 0. Fetch Settings
        try:
            rows = db.execute("SELECT key, value FROM settings").fetchall()
            settings = {row["key"]: row["value"] for row in rows}
        except Exception:
            settings = {}

        s_phone = settings.get("supplier_contact") or settings.get("supplier_phone", "")
        s_name = settings.get("supplier_name", "YOUR COMPANY")
        s_addr = settings.get("supplier_address", "")

        worksheet.write("A1", f"Tel. No. {s_phone}", fmt_tel)

        # Center across A-F
        worksheet.merge_range("A3:F3", s_name, fmt_company)
        worksheet.set_row(2, 25)

        worksheet.merge_range("A4:F4", "Specialist Manufacturers and Suppliers of Engineering Solutions", fmt_branding)
        worksheet.merge_range("A5:F5", s_addr, fmt_branding)

        worksheet.merge_range("A6:F6", "GUARANTEE CERTIFICATE", fmt_title)
        worksheet.set_row(5, 25)

        # --- Info Block (Rows 7, 8, 9) ---
        # Left: Buyer (Cols A-D merged)
        # Right: Labels (Col E), Values (Col F)

        b_name = header.get("consignee_name") or "The Purchase Manager"
        b_company = "M/S Partner Engineering PSU Ltd."
        b_location = "CITY NAME"

        # Row 7
        worksheet.merge_range("A7:D7", b_name, fmt_box)
        worksheet.write("E7", "GC No. & Dt.:", fmt_box)
        gc_val = f"{header.get('gc_number', header.get('dc_number', ''))} {fmt_date(header.get('gc_date', header.get('dc_date')))}"
        worksheet.write("F7", gc_val, fmt_box)

        # Row 8
        worksheet.merge_range("A8:D8", b_company, fmt_box)
        worksheet.write("E8", "PO No. & Dt.:", fmt_box)
        po_val = f"{header.get('po_number', '')} {fmt_date(header.get('po_date'))}"
        worksheet.write("F8", po_val, fmt_box)

        # Row 9
        worksheet.merge_range("A9:D9", b_location, fmt_box)
        worksheet.write("E9", "DC No. & Dt:", fmt_box)
        dc_val = f"{header.get('dc_number', '')} {fmt_date(header.get('dc_date'))}"
        worksheet.write("F9", dc_val, fmt_box)

        # --- Table Header (Row 10) ---
        worksheet.write("A10", "P.O.\nSl. No.", fmt_table_header)
        worksheet.merge_range("B10:E10", "Description", fmt_table_header)
        worksheet.write("F10", "Quantity", fmt_table_header)
        worksheet.set_row(9, 30)

        # --- Items ---
        curr = 10
        total_qty = 0
        for idx, item in enumerate(items):
            # Qty Logic
            qty_val = float(item.get("dsp_qty") or 0)
            total_qty += qty_val
            q_num = int(qty_val) if qty_val == int(qty_val) else qty_val
            unit = item.get("unit", "NOS")
            if str(unit).upper() == "NO":
                unit = "NOS"  # Normalize
            qty_str = f"{q_num} {unit}"

            # Desc
            desc = item.get("material_description") or item.get("description", "")

            # Calc Height
            lines = len(desc) // 50 + 1
            height = max(30, lines * 15)
            worksheet.set_row(curr, height)

            worksheet.write(curr, 0, item.get("po_item_no", idx + 1), fmt_cell_center)
            worksheet.merge_range(curr, 1, curr, 4, desc, fmt_cell_desc)
            worksheet.write(curr, 5, qty_str, fmt_cell_desc)  # Align Left as per image? Or Center? Image shows "11 NO" leftish.
            curr += 1

        # --- Empty Rows to Fill Visual Space (Full Grid as per Revert request) ---
        for _ in range(3):
            worksheet.write(curr, 0, "", fmt_cell_center)
            worksheet.merge_range(curr, 1, curr, 4, "", fmt_cell_desc)
            worksheet.write(curr, 5, "", fmt_cell_center)
            worksheet.set_row(curr, 25)
            curr += 1

        # --- GST Bill No Line (Requested from DC) ---
        inv_no = header.get("invoice_number", "")
        inv_dt = header.get("invoice_date")
        gst_line = f"GST Bill No. {inv_no}  Dt. {fmt_date(inv_dt) if inv_dt else ''}"

        # Border: Left, Right, Bottom (to close the table section before guarantee text)
        fmt_gst_line = workbook.add_format({"border": 1, "align": "left", "valign": "vcenter", "font_name": font_name, "font_size": 11, "indent": 1})

        # Write GST Line across columns (0-5)
        worksheet.merge_range(curr, 0, curr, 5, gst_line, fmt_gst_line)
        worksheet.set_row(curr, 25)
        curr += 1

        # --- Footer (Guarantee Text) ---
        footer_text = (
            "The goods supplied as above are guaranteed against manufacturing defects for 24 Month "
            "from delivery date. We undertake to replace or rectify the materials free of cost if any defects "
            "occur during this period."
        )
        worksheet.merge_range(curr, 0, curr + 2, 5, footer_text, fmt_footer_text)

        # --- Signature (No Box, Right Aligned) ---
        sig_r = curr + 4
        fmt_sig_plain = workbook.add_format({"bold": True, "font_name": font_name, "font_size": 11, "align": "right"})
        s_name = settings.get("supplier_name", settings.get("company_name", "YOUR COMPANY"))
        # Merge mostly to the right (Cols 3-5 = D-F)
        worksheet.merge_range(sig_r, 3, sig_r, 5, f"For {s_name}", fmt_sig_plain)

        workbook.close()
        output.seek(0)
        filename = f"GC_{header.get('dc_number', 'Draft')}.xlsx"
        return ExcelService._save_or_stream(output, filename, save_path)

    @staticmethod
    def _save_or_stream(output: io.BytesIO, filename: str, save_path: str = None):
        """
        Helper method to either save the BytesIO content to a file or stream it.
        """
        if save_path:
            try:
                # 1. Determine Full Path
                full_path = save_path
                # If it doesn't look like a file, treat as directory and append filename
                if not save_path.lower().endswith(".xlsx"):
                    full_path = os.path.join(save_path, filename)

                # 2. Ensure the PARENT directory of the FULL PATH exists
                # This works whether save_path was a file path or a directory path
                os.makedirs(os.path.dirname(full_path), exist_ok=True)

                # Check for duplications and increment filename if needed
                base, ext = os.path.splitext(full_path)
                counter = 1
                while os.path.exists(full_path):
                    full_path = f"{base}({counter}){ext}"
                    counter += 1

                with open(full_path, "wb") as f:
                    f.write(output.getvalue())
                logging.info(f"Excel file saved to: {full_path}")
                return {"success": True, "path": full_path, "saved_to_disk": True}
            except Exception as e:
                logging.error(f"Error saving to confirmed path {save_path}: {e}")

                # FALLBACK LOGIC: Windows Default Downloads
                try:
                    from pathlib import Path

                    fallback_dir = str(Path.home() / "Downloads")
                    os.makedirs(fallback_dir, exist_ok=True)

                    fallback_path = os.path.join(fallback_dir, filename)
                    # Deduplicate fallback path
                    base, ext = os.path.splitext(fallback_path)
                    counter = 1
                    while os.path.exists(fallback_path):
                        fallback_path = f"{base}({counter}){ext}"
                        counter += 1

                    with open(fallback_path, "wb") as f:
                        f.write(output.getvalue())

                    return {"success": True, "path": fallback_path, "saved_to_disk": True, "fallback_used": True}
                except Exception as fallback_err:
                    logging.error(f"Fallback save failed: {fallback_err}")
                    raise e  # Raise original error if fallback also dies
        else:
            headers = {"Content-Disposition": f'attachment; filename="{filename}"'}
            return StreamingResponse(
                output,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers=headers,
            )
