
import os

import openpyxl

# Add backend to path to find templates if needed
TEMPLATE_DIR = "backend/templates"

def wash_template(file_path):
    print(f"🧽 Washing template: {file_path}")
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    # Cells to clear (Supplier/Buyer Identity)
    cells_to_clear = ["A3", "A4", "A5", "A6", "A8", "A9", "A11", "A12", "C10", "K3", "O3", "I4", "K6", "O6", "K8", "O8", "Q4", "M10", "O10", "Q10"]
    
    for coord in cells_to_clear:
        try:
            cell = ws[coord]
            # Handle merged cells by writing to the top-left cell of the range
            from openpyxl.cell.cell import MergedCell
            if isinstance(cell, MergedCell):
                for m_range in ws.merged_cells.ranges:
                    if coord in m_range:
                        ws.cell(row=m_range.min_row, column=m_range.min_col).value = None
                        break
            else:
                cell.value = None
        except Exception as e:
            print(f"  ⚠️ Could not clear {coord}: {e}")

    # Special case: Payment Terms (Q4) - User wants "45"
    # Wait, the user said "save a blank Excel template with only field names".
    # So clearing Q4 is correct. The code will populate "45".

    # Sanitize "Net 45" strings everywhere in the sheet
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str) and "Net 45" in cell.value:
                cell.value = cell.value.replace("Net 45", "45")

    wb.save(file_path)
    print(f"✅ Template {os.path.basename(file_path)} washed successfully.")

def main():
    if not os.path.exists(TEMPLATE_DIR):
        print(f"❌ Template directory not found: {TEMPLATE_DIR}")
        return

    templates = [f for f in os.listdir(TEMPLATE_DIR) if f.endswith(".xlsx")]
    for t in templates:
        wash_template(os.path.join(TEMPLATE_DIR, t))

if __name__ == "__main__":
    main()
